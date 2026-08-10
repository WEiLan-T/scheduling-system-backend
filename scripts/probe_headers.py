# -*- coding: utf-8 -*-
import openpyxl

path = r'g:\ASOE\scheduling-system-backend\数据表\工艺路线数据大盘.xlsx'
wb = openpyxl.load_workbook(path)
out = []
out.append('sheetnames: ' + repr(wb.sheetnames))
ws = wb['全流水线工艺BOM']
out.append('max_row: %d max_col: %d' % (ws.max_row, ws.max_column))
out.append('merged_cells: ' + repr(list(ws.merged_cells.ranges)[:40]))
# 前3行全部单元格的原始repr
for r in range(1, 4):
    for c in range(1, 20):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            out.append(f'R{r}C{c}: repr={v!r}')
# A列前12个值与B列前12个值
out.append('--- col A/B/F samples ---')
for r in list(range(1, 13)) + [351, 352]:
    a = ws.cell(row=r, column=1).value
    b = ws.cell(row=r, column=2).value
    f = ws.cell(row=r, column=6).value
    out.append(f'row{r}: A={a!r} B={b!r} F={f!r}')
wb.close()
with open(r'g:\ASOE\scheduling-system-backend\target\header_probe.txt', 'w', encoding='utf-8') as fp:
    fp.write('\n'.join(out))
print('done')
